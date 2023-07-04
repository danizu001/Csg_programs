import pandas as pd
import excel as xls
from os import listdir
import datetime
from openpyxl import load_workbook
import os
from os.path import isfile, join, isdir

db_location = os.getcwd()+"\\Reports Database.xlsx"
bill_type_dic = {'SW':'\\Switched Billing','FA':'\\Facility Billing', 'RC':'Recip_Comp', '':''}

# ! need to check usage by day to confirm everything is okay
def renaming_reports (company,month,year,bill_type = ''):
    
    print('Going to rename files')
    number_of_files=0

    prefix_dict = pd.read_excel(db_location,sheet_name='Prefix', index_col=0,converters={'bill date':str}).to_dict()   
    path = "K:\\Client\\" + company + "\\Reports\\" + year + "\\" + month + "_" + year + bill_type_dic[bill_type]
   
    df = pd.read_excel(db_location,sheet_name=company)
    
    # # ! needs to be deleted after finish testing
    # path = 'K:\\Client\\GCI\\Reports\\2021\\05_2021\\Switched_Access\\test'


    before_renaming = tuple (df.iloc[:,0].dropna().tolist())
    after_renaming = tuple (df.iloc[:,1].dropna().tolist())
    
    if company == 'Neutral_Tandem' and bill_type == 'SW':
        not_OSA_TTS_files = tuple (df.iloc[:,5].dropna().tolist())

    code = pd.read_excel(db_location,usecols=[0,1,2,3,4],sheet_name=company,index_col=0,converters={'Code 1':str,'Code 2':str,'Code 3':str}).dropna().to_dict()

    files_list = [ f for f in listdir(path) if isfile(join(path, f))]

    for file in files_list:
        for report in before_renaming:
            index = before_renaming.index(report)
            if file.upper().startswith(str(report).upper()):
                extension = file.split('.')
                if company == 'Neutral_Tandem'and bill_type == 'SW':
                    if report in not_OSA_TTS_files:
                        file_name = prefix_dict['prefix'][company+bill_type] + '_' + month + year[2:] +'_'
                    elif extension[0].endswith('_02') or extension[0].endswith('_04'):
                        file_name = prefix_dict['prefix'][company+bill_type+'T'] + '_' + month +prefix_dict['bill date'][company+bill_type+'T'] + year[2:] +'_'
                    else:
                        file_name = prefix_dict['prefix'][company+bill_type+'O'] + '_' + month +prefix_dict['bill date'][company+bill_type+'O'] + year[2:] +'_'
                else:
                    file_name = prefix_dict['prefix'][company+bill_type] + '_' + month +prefix_dict['bill date'][company+bill_type] + year[2:] +'_'

                try:
                    if code['Code 1'][report] == '0':
                        code['Code 1'][report] = ''
                    os.rename(path+ "\\" +file, path+ "\\"+file_name + code['Code 1'][report].upper() +after_renaming[index]+'.'+extension[1])
                    number_of_files+=1
                except:
                    try:
                        if code['Code 2'][report] == '0':
                            code['Code 2'][report]= ''
                        os.rename(path+ "\\" +file, path+ "\\"+file_name + code['Code 2'][report].upper() +after_renaming[index]+'.'+extension[1])
                        number_of_files+=1
                    except:
                        if code['Code 3'][report] == '0':
                            code['Code 3'][report] = ''
                        os.rename(path+ "\\" +file, path+ "\\"+file_name + code['Code 3'][report].upper() +after_renaming[index]+'.'+extension[1])
                        number_of_files+=1
                print(file + ' Renamed')
                break   
    print(number_of_files,' files were renamed')

def convert_files(path):
    number_of_files=0
    not_needed_files_ends = ['AUR_History_Stats.xls','AUR History Stats Report.xls','AUR_History_Report.xls','Billing Notes.xls', 'Bill Count.xls','FGA_Usage_Balance_Details.xls']
    not_needed_files_starts = ['Fairpoint_moucomp','fusctax','Intrastate Terminating','Revenue_NECA','FF']
    filesList = [f for f in listdir(path) if isfile(join(path, f))]
    for file in filesList:
        if ((file.endswith(".xls")) and (file.endswith(tuple(not_needed_files_ends)) is False) and (file.startswith(tuple(not_needed_files_starts)) is False)):
            print("converting excel files " + file + " to XLSX version")
            xls.convert_xls_to_xlsx(path + "\\" + file)
            number_of_files+=1
    print(number_of_files,' files were converted')

def get_all_files(company,month,year,bill_type=''):
    

    path = 'K:\\Client\\' + company + '\\Reports\\' + year + "\\" + month + "_" + year + bill_type_dic[bill_type]


    print('Importing files location to database')

    # # ! needs to be deleted after finish testing
    # path = 'K:\\Client\\GCI\\Reports\\2021\\05_2021\\Switched_Access\\test'
    
    df_mps = pd.read_excel(db_location,sheet_name='mps_database',converters={'MPS Location':str,'Type':str,'Report Name':str})

    #getting the name of the files on the K regular report folder
    files_list = [file for file in listdir(path) if isfile(join(path, file))]


    #getting the code of all reports (including S2C) from the db and saving them into a list
    mps_code_list = df_mps['MPS Location'].values.tolist()
    report_name_list = df_mps['Report Name'].values.tolist()
    

    for file in files_list:  
        name=file.split("_")
        extension = name[-1].split('.')
        try:
            if 'OSA' in name or 'TTS' in name:
                if name[-2]+name[1] in mps_code_list :
                    index = mps_code_list.index(name[-2]+name[1])
                    df_mps.iloc[index,2] = path+'\\'+file

                elif extension[0]+name[1] in report_name_list:
                    index = report_name_list.index(extension[0]+name[1])
                    df_mps.iloc[index,2] = path+'\\'+file
            else:

                if name[-2] in mps_code_list and file.startswith('MPS') != True:
                    index = mps_code_list.index(name[-2])
                    df_mps.iloc[index,2] = path+'\\'+file
                elif file.startswith('MPS'):
                    index = mps_code_list.index('MPS')
                    df_mps.iloc[index,2] = path+'\\'+file
                elif extension[0] in report_name_list:
                    index = report_name_list.index(extension[0])
                    df_mps.iloc[index,2] = path+'\\'+file
        except:
            print(file,' is not named probably.',name)


    book = load_workbook(db_location)
    writer = pd.ExcelWriter(db_location, engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df_mps.to_excel(writer, "mps_database",index=False)
    writer.save()
    print('Importing files location to database is complete')

def retrieve_coordinates(code, option = ''):
    df = pd.read_excel(db_location, sheet_name = 'mps_database')

    mps_code_list = df['MPS Location'].values.tolist()

    if code+option in mps_code_list:
        index = mps_code_list.index(code+option)
        return(df['Coordinates'][index])

def retrieve_path(code, option = ''):
    df = pd.read_excel(db_location, sheet_name = 'mps_database')

    mps_code_list = df['MPS Location'].values.tolist()

    if code+option in mps_code_list:
        index = mps_code_list.index(code+option)
        return(df['File Location'][index])

def get_month():
    #dt = datetime.datetime.today()
    #month=str(dt.month)
    #year=str(dt.year)
    #if company == "SELECTRONICS":
        #month=month+1
        #month = month.zfill(2)
    month="06"
    year="2021"
    return(month,year)

def clean_database():
    #opening the MPS spreadsheet
    workbook = load_workbook(db_location)
    #opening the tab
    sheet = workbook['mps_database']
    
    #set all rows in column D as None
    for row in sheet.iter_rows (2, sheet.max_row, 3, 3):
        row[0].value = None
        
    #saving the spreadsheet
    workbook.save(db_location)
 
    #close the workbook after reading
    workbook.close()

