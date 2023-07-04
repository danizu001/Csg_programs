# this file contains all functions that we use to work with excel files.
from string import ascii_uppercase
import os, os.path
import win32com.client
import openpyxl
from openpyxl import  load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
import pandas as pd
from datetime import datetime
from datetime import timedelta
from dateutil.relativedelta import relativedelta
import shutil
import pdf_reports
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
import calendar
from openpyxl.worksheet.filters import (
    FilterColumn,
    CustomFilter,
    CustomFilters,
    DateGroupItem,
    Filters,
    )

mps_path =''


colors = {'red': 'F95E5E', 'blue': '5EC7F9', 'green': '00FF00'}

#Paste value in MPS cell. this functoin takes the input 'value' and cell location. the cell location is located in our database. and it take a comment in case that there is a comment to write.
def paste_MPS(input, cell, comment = 0):    
    #opening the MPS spreadsheet and selecting the main sheet
    MPS_workbook = load_workbook(mps_path)
    MPS_sheet = MPS_workbook.active
    
    #pasting the value in a specific cell
    MPS_sheet[cell] = input

    # adding comments in case there was any comments
    if comment !=0:
        comment = Comment(comment, "CABS Python")
        MPS_sheet[cell].comment = comment
    #saving the spreadsheet
    MPS_workbook.save(mps_path)

    #close the workbook after reading
    MPS_workbook.close()
    



#Autosize the width of columns. it is used to format excel reports. this function takes the report path.
def autosize_columns(report_path):
    #opening the report spreadsheet and selecting the main sheet
    report_workbook = load_workbook(report_path)
    report_sheet = report_workbook.active

    #list to save the max len of each column
    column_max_len = []
    
    #checking how many column and setting the list width accordingly. 
    for column in range(report_sheet.max_column):       
        column_max_len += [0]

    for row in report_sheet:        #it will loop in all rows till reach the limit_row in the active sheet          
        #enumerate will assign an index i to each element of row (the elements will be the columns)
        for column_index, cell in enumerate(row):            
            #check if the current len of the element is smaller, if so it replaces to get the longest row.  
            if len(str(cell.value)) > column_max_len[column_index]:
                column_max_len[column_index] = len(str(cell.value))
        
    for i, column_width in enumerate(column_max_len):
        #adjust the width of the column according to the max len of each one 
        report_sheet.column_dimensions[get_column_letter(i+1)].width = column_width+3
  

    #saving the spreadsheet
    report_workbook.save(report_path)
    
    #close the workbook after reading
    report_workbook.close()




#Convert xls file to xlsx and delete the old xls file
#NOTES:
#- Applying converters reading the excel file will allow to save 0069 FCCID as 0069 instead of just 69
#- Converting dates to str won't work well since it does not allow to deal filters as date format does
def convert_xls_to_xlsx(report_path):
    xls_file = pd.read_excel(report_path,
        converters={'bill_fccid':str, 'ic_cic':str, 'clli_co':str, 'user_invoice_num':str, 'invoice_num':str})       #read with Pandas
    new_path = report_path.replace('.xls', '.xlsx')        #creating the new path for .xlsx file
    xls_file.to_excel(new_path, index = False)      #converting the original file to xlsx
    os.remove(report_path)     #removing the old file



#Calculate the total of one column
#if there is no need for filtering, just don't enter any information in bill_type nor bill_date
#if you want to filter by bill type, enter the bill type as 'FA','SW' or 'RC' 
#if you want to filter by bill date, use the format 'YYYY-MM-DD' in single quotes 
#NOTES: Not usable for Deletion reports or others where specific filtering is required  
def sum_column(report_df, name_sum_column, bill_type = 'N/A', bill_date = 'N/A',disconnect='N'):
    #read the report
    df = report_df
    
    #SW tuple that contains initial characters in user invoice num column - we will use it in the bill type condition
    SW_tup = ('D', 'A0', 'A1', 'A2', 'A3', 'A4', 'A5', 'A6', 'A7', 'A8', 'A9')

    #check the bill_type
    if bill_type.upper() == 'SW':

        # We use try and catch due to reports structure. Not all reports has user invoice number.
        try:
            #filtering under SW invoices
            df = df[['user_invoice_num', name_sum_column]].where(df['user_invoice_num'].str.startswith(SW_tup)).dropna()
        except:
            df = df[['ban', name_sum_column]].where(~df['ban'].str.contains('FA')).dropna()
            df = df[['ban', name_sum_column]].where(~df['ban'].str.contains('O')).dropna() #~ as negation, it will display the rows where BAN column does not contains 'FA'
    elif bill_type.upper() == 'FA':
        try:
            #filtering under FA invoices
            df = df[['user_invoice_num', name_sum_column]].where(~df['user_invoice_num'].str.startswith(SW_tup)).dropna()
            df = df[['user_invoice_num', name_sum_column]].where(~df['user_invoice_num'].str.startswith('O')).dropna()      #we use ~ as negation
        except:
            df = df[['ban', name_sum_column]].where(df['ban'].str.contains('FA')).dropna() #it will display the rows where BAN column contains 'FA'
    elif bill_type.upper() == 'RC':
        df = df[['ban', name_sum_column]].where(df['ban'].str.contains('O')).dropna()

    #check the bill_date
    if bill_date != 'N/A':
        df = df[['bill_date', name_sum_column]].where(df['bill_date'] == bill_date).dropna()

    # For FA Billed with Disconnects we filter by 'Disconnected' circuits and calculate the rev for them
    if disconnect == 'Y':
        df = df[['status', name_sum_column]].where(df['status'] == 'Disconnected').dropna()


    #sum column values
    value = round(df[name_sum_column].sum(),2)
    
    return value


#Calculate the total of many columns 
#It works with a range of columns, they need to be next to each other 
#Enter just the initial and the last column in the range you want to sum
#Use column names, not numbers nor letters
def sum_columns_range(report_df, initial_column, last_column):
    #read the report
    df = report_df

    #select all rows for the columns we need
    df = df.loc[:,initial_column:last_column]

    #sum each column
    sum_column = df.sum(axis = 0, skipna = True)

    #sum the totals of each column 
    total = sum_column.sum().round(2)

    return total


#Add filter
#This function will create a filter condition inside excel files
#It is usefull because it keeps all the data outside the condition. Something we cannot do with Pandas
#It shows efficiency in small-size reports, not for hundred thousand rows spreadsheets
#In column filter enter one upper letter in single quotes as 'A' or 'Z'

#PROBLEMS:
#Openpyxl stablishes the filters but it cannot apply them
#In Pandas we can filter but we will delete all rows that do not match the condition
#There is a possible solution using xlswriter (this function is not needed anymore so we will not try using xlswriter)

def add_filter(df, name_filter_column, values_to_filter):#values_to_filter is a list that could contain another list inside
    for i in range(len(values_to_filter)):
        try:
            df_filter=df[df[name_filter_column[i]]==values_to_filter[i]]
        except:
            df_filter=df[df[name_filter_column[i]].isin(values_to_filter[i])]
        df=df_filter
    return df_filter

def get_total_days():
    #Getting the 'from' and 'to cycle' dates from the MPS
    dates = get_values_from_excel (['D22', 'D23'], mps_path ,  sheet=0)
    from_cycle_date = dates[0]
    to_cycle_date = dates[1]
    prev_from=from_cycle_date+relativedelta(months=-1)
    prev_to=to_cycle_date+relativedelta(months=-1)
    #Calculating the amount of days between from and to cycle dates
    #start_date = datetime.strptime(from_cycle_date, "%m/%d/%Y")
    #end_date = datetime.strptime(to_cycle_date, "%m/%d/%Y")
    current_days = (to_cycle_date-from_cycle_date).days + 1       #adding 1 to include the starting date
    prev_days = (prev_to-prev_from).days + 1       #adding 1 to include the starting date
    #Setting the dates between from and to cycle dates
    date_range = []
    for day in range(current_days):
        date_range.append(from_cycle_date + timedelta(day))
    return date_range,current_days,prev_days



#Verify all days are included in usage by day report
#This function will verify if all days in a month have data 
#it takes the range date from the MPS
def verify_days(report_df):

    date_range,current_days,prev_days=get_total_days()

    #Getting unique FCCIDs
    df = report_df
    FCCIDs = df['bill_fccid'].unique()

    #Checking if all dates between from and to cycle dates are in each FCCID
    missing_dates = []      #missing dates will be saved here
    for FCCID in FCCIDs:
        print('Analizing ' + str(FCCID) + ' FCCID')
        for date in date_range:
            if not df['date_of_record'].where(df['bill_fccid'] == FCCID).dropna().isin([date]).any():       #any return the boolean
                missing_dates.append(date.strftime('%m/%d/%Y'))
    
    #Setting what to return
    #In case all dates are included
    if len(missing_dates) == 0:
        return True, date_range
    #In case there are missed dates
    else:
        return missing_dates, date_range



# to get specific value from an excel (mainly the MPS). the fn takes list of needed cells to return the values
# the report location and the sheet name in case it is not the default.
def get_values_from_excel (list_cells, report_location , sheet=0):

    value_cells = []
    #opening the report spreadsheet and selecting the sheet
    MPS_workbook = load_workbook(report_location)
    if sheet== 0:
        MPS_sheet = MPS_workbook.active
    else:
        MPS_sheet = MPS_workbook[sheet]

    #saving the values in a list "value_cells"
    for i in range(len(list_cells)):
        value_cells.append(MPS_sheet[list_cells[i]].value)
    
    return value_cells



# the fn counts the number of rows in an excel file. the fn takes report dataframe, the name of the column to count and param is a value in case we need to filter
def count_rows(report_df, name_count_column, param='N/A'):

    # apply the filter in case the is a filter and return the number of rows. else return the number of row for the given column.
    if param != 'N/A':
        count=report_df[name_count_column].value_counts()[param]
    else:
        count=report_df[name_count_column].count()
    return count



# this fn adds a new sheet to an existing excel. name_sheet and dataframes are lists. number of original sheets is the number of exising sheets before adding a new sheet.
def add_sheet(report_location, name_sheet, dataframes, number_original_sheets, usg_by_date=False): 

    # changing the writer to xlsxwriter
    writer = pd.ExcelWriter(report_location, engine = 'xlsxwriter')

    # saving all sheets to a new excel (overwriting the existing one)
    for i in range(len(name_sheet)):

        # in case that it the original sheet or it is usage by day report we use index false
        if i<=number_original_sheets-1 :
            dataframes[i].to_excel(writer, sheet_name = name_sheet[i],index=False)

        # in case of the new sheet we use index true
        elif i>number_original_sheets-1 : 
            dataframes[i].to_excel(writer, sheet_name = name_sheet[i])
        
    writer.save()



# this function creates a pivot table based on date range and fccid.
def create_pivot_table_fccid_range_date(dataframe, date_range, pivot_column, pivot_index, pivot_value, function):

    #if condition to find which one is the fccid (pivot_column or pivot_index)
    if 'fccid' in pivot_column:
        df_pivot = dataframe[dataframe[pivot_index].isin(date_range)]
        pivot = df_pivot.pivot_table(values =[pivot_value], index =[pivot_index], columns = [pivot_column], aggfunc =function)
    else:
        df_pivot = dataframe[dataframe[pivot_column].isin(date_range)]
        pivot = df_pivot.pivot_table(index =[pivot_index],values =[pivot_value],aggfunc =function)
    
    return pivot



def excel_alphabet(fccids_number): #Is a function for have the alphabet in a excel form for call a range of cells 
    letters = list(ascii_uppercase)
    num_cols = fccids_number

    excel_cols = []
    for i in range(0, num_cols - 1):
        n = i//26
        m = n//26
        i-=n*26
        n-=m*26
        col = letters[m-1]+letters[n-1]+letters[i] if m>0 else letters[n-1]+letters[i] if n>0 else letters[i]
        excel_cols.append(col)
    return(excel_cols)



def change_color_by_list(report_location, cell, color , sheet='N/A'): #Change color in a range of a list
    wb = load_workbook(report_location)
    if sheet != 'N/A':
        MPS_sheet = wb[sheet]
    else:
        MPS_sheet = wb.active
    
    for i in range(len(cell)):
        MPS_sheet[cell[i]].fill = PatternFill(start_color=colors[color[i]], end_color=colors[color[i]], fill_type = "solid")  


    #saving the spreadsheet
    wb.save(report_location)
    

def hide_column(report_location, name_sheet, list_columns):
    wb=load_workbook(report_location)
    worksheet = wb[name_sheet]
    for col in list_columns:
        worksheet.column_dimensions[col].hidden= True
    wb.save(report_location)
    wb.close()    
    
def hide_row(report_location, name_sheet, list_rows):
    wb=load_workbook(report_location)
    worksheet = wb[name_sheet]
    for row in list_rows:
        worksheet.row_dimensions[row].hidden= True
    wb.save(report_location)
    wb.close()        
 
#Requirements of GCI for the sbt and fbt
def sbt_fbt_gci(path_in,path_out,lim,report,occ_path):
    report_df = pd.read_excel(path_in)
    for i in range(1,len(report_df)): #switch
        report_df.iloc[i,2:lim]=report_df.iloc[i,2:lim].astype(float)
    #Create the new columns and extract the values from the original sheet
    data=report_df[['ban','current_revenue','minus_01_revenue','minus_02_revenue','minus_03_revenue','minus_06_revenue']]
    data['OCC']=0.0
    data['current vs last month']=data['current_revenue'][1:]-data['minus_01_revenue'][1:]
    data['current vs 2 month ago']=data['current_revenue'][1:]-data['minus_02_revenue'][1:]
    data['current vs 3 month ago']=data['current_revenue'][1:]-data['minus_03_revenue'][1:]
    data['current vs 6 month ago']=data['current_revenue'][1:]-data['minus_06_revenue'][1:]
    if '12P' in occ_path:#Switch
        report_occ=pdf_reports.sbt_gci(occ_path)
        data_occ=report_occ.groupby(["ban"]).sum()['amount_billed']
        for i in range(len(data['ban'])):
            if data['ban'][i] in data_occ.index:
                data['OCC'][i]=round(float(data_occ[data['ban'][i]]),2)
    else:
        #Is only for fbt and put the number of the OCC billed in the new dataframe required
        report_occ = pd.read_excel(occ_path)
        data_occ=report_occ.groupby(["ban"]).sum()['amount_billed']
        for i in range(len(data['ban'])):
            if data['ban'][i] in data_occ.index:
                data['OCC'][i]=round(data_occ[data['ban'][i]],2)
    data=data[['ban','OCC','current_revenue','minus_01_revenue','minus_02_revenue','minus_03_revenue','minus_06_revenue','current vs last month','current vs 2 month ago','current vs 3 month ago','current vs 6 month ago']]
    new_row=['','','actual_date','last month','2 month ago','3 month ago','6 month ago','','','','']
    tail=data[0:]
    head = pd.DataFrame([new_row], columns=data.columns.values.tolist())
    data=head.append(tail)
    writer = pd.ExcelWriter(path_out)
    report_df.to_excel(writer, report, index=False)
    data.to_excel(writer,'data', index=False)
    writer.save()