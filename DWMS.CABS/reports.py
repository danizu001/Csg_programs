import re
import sys
import database as db
import excel as xls
import pdf_reports as pdf
from openpyxl import  comments, load_workbook
from os.path import isfile, join
from os import listdir
import numpy as np
import frontend
import os, os.path
import win32com.client
import pandas as pd
import calendar
import datetime
from numpy import nan
import holidays               
from datetime import datetime
import matplotlib.pyplot as plt

mps_path = xls.mps_path
temp_month=0
temp_year=0
macro_path = ''


def usage_by_day_report(code,option =''):
    
    day_type=[]
    average_wd=[]
    average_we=[]
    desv_est_wd=[]
    desv_est_we=[]
    holiday=[]

    try:
        
        report_location = db.retrieve_path(code,option)
        report_df = pd.read_excel(report_location, converters={'bill_fccid':str}).dropna(subset=['bill_fccid'])
        if code == '5D':
            # Making sure that there is usage for every day in the bill period      
            verify_days_result, date_range = xls.verify_days(report_df)

            # If all days are included. value will equal to the sum of mou
            if verify_days_result == True:
                value = xls.sum_column(report_df, "mou")
                # create a pivot table for each fccid with the total of mou per day and replace any empty value with zero
                df_pivot = xls.create_pivot_table_fccid_range_date(report_df, date_range, "bill_fccid", "date_of_record", "mou", 'sum')
                df_pivot = df_pivot.fillna(0)

                # getting all the holidays for the US
                us_holidays = holidays.UnitedStates()
                
                # day_type is to detrmine if the day is weekend or weekday. holiday is a list that will include the holidays for the bill period month.
                for day in df_pivot.index: # is the range of days
                    if(day in us_holidays): #Verify if the day is a holiday in US
                        day_type.append(True)
                        holiday.append(day)
                    else:    
                        day_type.append(day.weekday() > 4)
                
                df_pivot['weekend']=day_type# add weeked column to the pivot table.
                FCCIDs=df_pivot.columns[:-1] # is a list of the FCCIDs

                # calculating the mean and std for each fccid (for weekend days and for weekdays)
                for fccid in FCCIDs:
                    values_we=df_pivot[df_pivot['weekend']==True][fccid]
                    values_wd=df_pivot[df_pivot['weekend']==False][fccid]
                    average_wd.append(round(np.mean(values_wd),2))
                    average_we.append(round(np.mean(values_we),2))
                    desv_est_wd.append(round(np.std(values_wd),2))
                    desv_est_we.append(round(np.std(values_we),2))

                # adding an empty cells to match the number of row of the pivot table
                average_wd.append(nan)
                average_we.append(nan)
                desv_est_wd.append(nan)
                desv_est_we.append(nan)

                # adding the mean and std to the pivot table
                df_pivot.loc['Weekday Average']=average_wd
                df_pivot.loc['Weekend Average']=average_we
                df_pivot.loc['Standard deviation Weekday']=desv_est_wd
                df_pivot.loc['Standard deviation Weekend']=desv_est_we
                
                # saving the orignal report and the pivot table.
                sheet_name=['UsageByDay','statistics']
                data=[report_df,df_pivot]
                xls.add_sheet(report_location,sheet_name,data,1,True)#Change the 1 or 0 for true or fals in the weekend column
                
                df_stadistic=df_pivot.droplevel(0,axis='columns').iloc[:-4,:-1]
                df_stadistic2=[i.strftime('%m-%d') for i in df_stadistic.index]
                column_sum=[df_stadistic.iloc[i].sum() for i in range(len(df_stadistic))]

                # the part below is for showing a graph and save it as pdf          
                # fig = plt.figure()
                # ax = fig.add_axes([0,0,1,1])  
                # ax.bar(df_stadistic2 ,column_sum)
                # plt.xticks(rotation=75)
                # plt.title('Sum of mou all FCCID 2021')
                # plt.xlabel('Days of the month')
                # plt.ylabel('Value of mou')
                # plt.savefig('Plot_customer.pdf')
                # plt.show()
                
                columns_number = len(FCCIDs) + 2 # the number of columns 
                excel_columns_alphabet_name = xls.excel_alphabet(columns_number) # A list that will contain the columns name (columns name is alaphabet (A, B, C, etc ...))
                days_number = len(date_range) # the number of days for the bill cycle 

                # reading the pivot table for coloring
                wb = load_workbook(report_location, data_only=True)
                sh = wb["statistics"]
                cell = []
                color = []
                
                # selecting the holidays to change the color to green
                for date in range(days_number):
                    if df_pivot.index[date] in holiday:
                        cell.append('A'+str(date+4))
                        color.append('green')

                # coloring the cells based on if they are above (blue) or they are brlow (red)
                for alphabet in range(1,len(excel_columns_alphabet_name)):
                    for number in range(4,days_number) :
                        if ((sh[str(excel_columns_alphabet_name[alphabet])+str(number)].value < (sh[str(excel_columns_alphabet_name[alphabet])+str(days_number+4)].value-sh[str(excel_columns_alphabet_name[alphabet])+str(days_number+6)].value)) and (day_type[number-4]==False)):
                            cell.append(str(excel_columns_alphabet_name[alphabet])+str(number))
                            color.append('red')

                        if ((sh[str(excel_columns_alphabet_name[alphabet])+str(number)].value > (sh[str(excel_columns_alphabet_name[alphabet])+str(days_number+4)].value+sh[str(excel_columns_alphabet_name[alphabet])+str(days_number+6)].value)) and (day_type[number-4]==False)):
                            cell.append(str(excel_columns_alphabet_name[alphabet])+str(number))
                            color.append('blue')

                        if ((sh[str(excel_columns_alphabet_name[alphabet])+str(number)].value < (sh[str(excel_columns_alphabet_name[alphabet])+str(days_number+5)].value-sh[str(excel_columns_alphabet_name[alphabet])+str(days_number+7)].value)) and (day_type[number-4]==True)):
                            cell.append(str(excel_columns_alphabet_name[alphabet])+str(number))
                            color.append('red')

                        if ((sh[str(excel_columns_alphabet_name[alphabet])+str(number)].value > (sh[str(excel_columns_alphabet_name[alphabet])+str(days_number+5)].value+sh[str(excel_columns_alphabet_name[alphabet])+str(days_number+7)].value)) and (day_type[number-4]==True)):
                            cell.append(str(excel_columns_alphabet_name[alphabet])+str(number))
                            color.append('blue')
                xls.change_color_by_list(report_location, cell, color,sheet='statistics')
            # If there is any missing days. value will equal to the missing days and the sum of mou as well
            elif verify_days_result != True:
                value = str(verify_days_result) + " value = " +str(xls.sum_column(report_df, "mou"))

            
        else :
            value = xls.sum_column(report_df, "mou")
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS(value, cell)
        if code == "5E-3MONTHS":
            print("\033[1;32;40m" + sys._getframe().f_code.co_name + " 3 months is done" + "\033[0;37;40m")
            db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " 3 months is done\n")
        else:
            print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
            db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")            
    except :
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")
        

def usage_balancing_report(code, bill_type="NA",option=''):
    try:
        report_location = db.retrieve_path(code,option)
        value = pdf.usage_balancing(report_location,bill_type)
        cell = db.retrieve_coordinates(code,option)   
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
    except :
        cell = db.retrieve_coordinates(code,option) 
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")
        
        

def switched_bans_trending_report(code="12B",option ='',company=0):
    try:
        report_location = db.retrieve_path(code,option)
        if report_location.endswith('pdf'):
            value = pdf.bans_trending_report(report_location)
            comment = 0
        elif report_location.endswith('xlsx'):
            report_df = pd.read_excel(report_location)
            if company == 'gci':
                report_df= report_df.drop(labels=0, axis=0).astype({'current_revenue':'float', 'minus_01_revenue':'float', 'minus_01_minutes':'float', 'current_minutes':'float'})
                report_df = report_df.rename(columns = {'minus_01_minutes': 'minus_01_rated_minutes', 'current_minutes' : 'current_rated_minutes'})
            exists = 'NO BILL' in report_df['current_bill']
            if exists == True:
                missing_invoices = report_df['current_bill'].where(report_df['current_bill'].str.startswith('NO BILL')).dropna().count()
            else:
                missing_invoices=0                      
            if missing_invoices == 0:
                value = xls.sum_column(report_df, 'current_revenue')
                old_rev=xls.sum_column(report_df, 'minus_01_revenue')
                #Reviewing old Revenue against Current
                date_range,current_days,prev_days=xls.get_total_days()
                difference=(((value/current_days) / (old_rev/prev_days)) - 1 ) * 100
                if old_rev>value:
                    old_mou=report_df['minus_01_rated_minutes'].sum()
                    new_mou=report_df['current_rated_minutes'].sum()
                    if old_mou>new_mou:
                        comment='The Revenue and MOU are less than last month, old revenue:  '+str(old_rev)+"\nDifference per Day: "+str(round(difference,2))+'%' 
                    else:
                        comment='The Revenue is less than last month but the MOU is higher than last month, old revenue: '+str(old_rev)+" Old MOU: "+str(old_mou) +" Current MOU: "+str(new_mou)+"\nDifference per Day: "+str(round(difference,2))+'%' 
                elif old_rev<value:
                    comment='Revenue is higher than last month, old revenue: '+ str(old_rev)+"\nDifference per Day: "+str(round(difference,2))+'%'
                else:
                    comment='Please check the report, is equal to last month, old revenue: '
            else:
                value = (f'{missing_invoices} invoices not produced')
        if(company=='gci'):
             occ_path=report_location.split('\\')
             occ_file=occ_path[7].split('_')
             occ_file[3]='12P'
             occ_file[4]='OCC Report.pdf'
             occ_path[7]='_'.join(occ_file)
             occ_path='\\'.join(occ_path)
             xls.sbt_fbt_gci(report_location,report_location,54,'sbt',occ_path)
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS(value, cell,comment)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
    except :
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")
        
        

def facility_bans_trending_report(code="12B",bill_type='NA',option ='',company=0):
    try:
        report_location = db.retrieve_path(code,option)
        report_df = pd.read_excel(report_location)
        if company == 'gci':
            report_df= report_df.drop(labels=0, axis=0).astype({'current_revenue':'float', 'minus_01_revenue': 'float','current_available_ckts':'int', 'minus_01_available_ckts': 'int'})
        
        exists = 'NO BILL' in report_df['current_bill']
        if exists == True:
            missing_invoices = report_df['current_bill'].where(report_df['current_bill'].str.startswith('NO BILL')).dropna().count()
        else:
            missing_invoices=0


        if missing_invoices == 0:
            # getting (number of circuits and revenue) values for previous month and current month
            value = xls.sum_column(report_df, 'current_revenue')
            prev_month_value =  xls.sum_column(report_df, 'minus_01_revenue')
            number_circuits = xls.sum_column(report_df, 'current_available_ckts')
            prev_month_number_circuits = xls.sum_column(report_df, 'minus_01_available_ckts')


            comment = 'Revenue changed by : {:.2f}%\n'.format(((value/prev_month_value) - 1) * 100)
            if number_circuits > prev_month_number_circuits:
                comment = comment + 'Number of circuits increased from {} to {}\nDifference : {} circuits'.format(prev_month_number_circuits,number_circuits,number_circuits - prev_month_number_circuits)
            else : 
                comment = comment + 'Number of circuits decreased from {} to {}\nDifference : {} circuits'.format(prev_month_number_circuits,number_circuits,prev_month_number_circuits - number_circuits)
        if company=='gci':
             occ_path=report_location.split('\\')
             occ_file=occ_path[7].split('_')
             occ_file[3]='19BK'
             occ_file[4]='OCC Billed Report.xlsx'
             occ_path[7]='_'.join(occ_file)
             occ_path='\\'.join(occ_path)
             xls.sbt_fbt_gci(report_location,report_location,28,'fbt',occ_path)
        else:
            value = (f'{missing_invoices} invoices not produced')
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS(value, cell,comment)
        cell = db.retrieve_coordinates('19AG',option)
        xls.paste_MPS(number_circuits, cell)

        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
    except :
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS("file is not found", cell)
        cell = db.retrieve_coordinates('19AG',option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")
        
        

def adjustment_report(code='12O', option =''):
    try:
        report_location = db.retrieve_path(code,option)

        if report_location.endswith(".xlsx"):
            report_df = pd.read_excel(report_location)
            value = xls.sum_column(report_df, "amount_billed")
        else:
            value = pdf.adjustment(report_location)
        cell = db.retrieve_coordinates(code, option)
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
        
        
    except :
        cell = db.retrieve_coordinates(code, option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")
        
        

def occ_report(code='12P', bill_type = 'N/A',option =''):
    try:
            
        report_location = db.retrieve_path(code,option)

        if report_location.endswith(".xlsx"):
            report_df = pd.read_excel(report_location)
            value = xls.sum_column(report_df, "amount_billed",bill_type)
        else:
            value = pdf.occ(report_location)
        cell = db.retrieve_coordinates(code, option)
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
        
        
    except :
        cell = db.retrieve_coordinates(code, option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")
        
        

def payment_report(fa_sw_rc="N/A", date="N/A",option =''):
    try:
            
        report_location = db.retrieve_path("8",option)
        
        if report_location.endswith(".xlsx"):
            report_df = pd.read_excel(report_location)
            value = xls.sum_column(report_df, "payment_amount", bill_type=fa_sw_rc, bill_date=date)
        if report_location.endswith(".pdf"):
            value = pdf.payment(report_location)

        cell = db.retrieve_coordinates("8",option)
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
        
    except :
        cell = db.retrieve_coordinates("8",option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")
        
        

def aged_trial_balance_report(option =''):
    try:

        report_location = db.retrieve_path("19A",option)
        value = pdf.aged_trial_balance(report_location)
        cell = db.retrieve_coordinates("19A",option)
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
        
        
    except :
        cell = db.retrieve_coordinates("19A",option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")       
        

def aged_trial_balance_for_export_report(code="19B",bill_type='NA',option =''):
    try:
        report_location = db.retrieve_path(code,option)
        report_df = pd.read_excel(report_location, converters={'bill_fccid':str, 'ic_cic':str}).dropna(subset=['bill_fccid'])
        
        if code == "19B":
            value = xls.sum_column(report_df, "compute_0010",bill_type)
        if code == "13B":
            value = xls.sum_column(report_df, "compute_0010", bill_date=load_workbook(xls.mps_path).active['D24'].value)


        sheet_names=["Aged Trial Balance for Export","pivot_table"]

        if  code=='13B': 
            cell=['D24']
            bill_date=xls.get_values_from_excel(cell,xls.mps_path)
            df_pivot=xls.create_pivot_table_fccid_range_date(report_df,bill_date,"bill_date","bill_fccid","compute_0010",'sum')#0 to show the pivot table in vertical
            data=[report_df,df_pivot]
            xls.add_sheet(report_location,sheet_names,data,1)#The one means how many original Tabs had the file before any changes

        elif code=='19B':
            
            FCCIDs = report_df['bill_fccid'].unique()
            df_pivot = report_df[report_df['bill_fccid'].isin(FCCIDs)]
            df_pivot = report_df.pivot_table(index =['bill_fccid'],values =['compute_0010'],aggfunc ='sum')
            data=[report_df,df_pivot]
            xls.add_sheet(report_location,sheet_names,data,1)
        

        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
        
        
    except :
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")
        
        

def mmr_by_BAN_report(code="19G",option =''):
    try:
            
        report_location = db.retrieve_path(code,option)        
        value = pdf.mmr(report_location)
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
        
        
    except :
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")
        
        

def transaction_summary_report(code="19M",client='all',bill_type="NA",option =''):
    try:
        report_location = db.retrieve_path(code,option) 
        if report_location.endswith('.xlsx'):
            report_location= report_location.replace('.xlsx', '.pdf')       
        value = pdf.transaction_summary(report_location,bill_type)
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
        
        
        if client == 'peerless' :
            value = pdf.late_payment_charge_from_TS(report_location)
            cell = db.retrieve_coordinates("19Q",option)
            xls.paste_MPS(value, cell)
            print("late_payment_charges_report is done")
        if client == 'NT' :
            report_location = report_location.replace('.pdf', '.xlsx')
            report_df = pd.read_excel(report_location)
            report_df=report_df[report_df["name"]=="Billing & Collection Revenue"]
            value=round(report_df["compute_0005"].sum(),2)
            cell = db.retrieve_coordinates("19Q",option)
            new=report_location.replace('19M_', '')
            os.rename(report_location,new)
            xls.paste_MPS(value, cell)
            print("late_payment_charges_report is done")

    except :
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")
        
        

# needs to be checked
def late_payment_charges_report(option =''):
    try:
        report_location = db.retrieve_path("19Q",option)
        report_df = pd.read_excel(report_location)
        cell = db.retrieve_coordinates("19Q",option)
        value = xls.sum_columns_range(report_df, "is_lpc1", "lc_lpc1")
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
        
        
    except :
        cell = db.retrieve_coordinates("19Q",option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")
        
        

def revenue_analysis_report(code = '19S', option ='', bill_type = 'N/A'):
    try:

        report_location = db.retrieve_path(code,option)
        report_df = pd.read_excel(report_location)
        value = xls.sum_column(report_df, "monthly_revenue", bill_type = 'N/A')
        cell = db.retrieve_coordinates(code, option)
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
        
        
    except :
        cell = db.retrieve_coordinates("19S",option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")
        
        

def bill_by_period_report(option =''):
    try:
        report_location = db.retrieve_path("19V",option)
        report_df = pd.read_excel(report_location)
        try:
            value = xls.sum_column(report_df, "total")
        except:
            value=xls.sum_columns_range(report_df,"current_pd","pd_three") 
        cell = db.retrieve_coordinates("19V",option)
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
        
        
    except :
        cell = db.retrieve_coordinates("19V",option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")
        
        

def facility_circuit_charges_billed_disconnected(code="19AF",disconnect='N',option =''):
    try:
        report_location = db.retrieve_path(code,option)
        report_df = pd.read_excel(report_location)

        value = xls.sum_column(report_df, "occ_rev",  disconnect=disconnect)
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
        
        
    except :
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")
        
        

def invoice_balance_report(option =''):
    try:
        report_location = db.retrieve_path("19AL",option)
        report_df = pd.read_excel(report_location)
        try:
            value = xls.sum_column(report_df, "compute_0008")
        except KeyError:
            try:
                value = xls.sum_column(report_df, "balance")
            except:
                value = xls.sum_column(report_df, "tot_amt")
        cell = db.retrieve_coordinates("19AL",option)
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
        
        
    except :
        cell = db.retrieve_coordinates("19AL",option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")
        
        

def msg_mou_rev_with_lpc(code="19E", option =''):
    try:
        report_location = db.retrieve_path(code,option)
        report_df = pd.read_excel(report_location)
        value = xls.sum_column(report_df, "cur_rev1")
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
        
        
    except :
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")
        
        

# needs to check "I dont know what to check"
def msg_mou_rev_export(bill_type='NA',option =''):
    try:

        report_location = db.retrieve_path("19F",option)
        report_df = pd.read_excel(report_location)
        if bill_type == 'FA':
            value = xls.sum_columns_range(report_df, "occ_iser_rev1", "occ_local_rev1")
        else:
            value = xls.sum_columns_range(report_df, "usg_iser_rev1", "usg_local_rev1")
        cell = db.retrieve_coordinates("19F",option)
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
        
        
    except :
        cell = db.retrieve_coordinates("19F",option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")
        
        

def mmr_by_clli_report(option =''):
    try:
            
        report_location = db.retrieve_path("19H",option)
        value = pdf.mmr(report_location)
        cell = db.retrieve_coordinates("19H",option)
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
        
        
    except :
        cell = db.retrieve_coordinates("19H",option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")
        
        

def switched_usage_summary_charges(option =''):
    try:
            
        report_location = db.retrieve_path("19X",option)
        report_df = pd.read_excel(report_location)

        value = xls.sum_column(report_df, "compute_0016")
        cell = db.retrieve_coordinates("19X",option)
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
        
        
    except :
        cell = db.retrieve_coordinates("19X",option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")
        
        

def facility_summary_charges(option =''):
    try:

        report_location = db.retrieve_path("19AB",option)
        report_df = pd.read_excel(report_location)

        value = xls.sum_column(report_df, "compute_0012")
        cell = db.retrieve_coordinates("19AB",option)
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
        
        
    except :
        cell = db.retrieve_coordinates("19AB",option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")
        
        

def facility_charges_by_cic_clli(option =''):
    try:
            
        report_location = db.retrieve_path("19AD",option)
        report_df = pd.read_excel(report_location)

        value = xls.sum_column(report_df, "compute_0008")
        cell = db.retrieve_coordinates("19AD",option)
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
        
        
    except :
        cell = db.retrieve_coordinates("19AD",option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")
        
        

def fusc_charge_by_circuit(option =''):
    try:
        report_location = db.retrieve_path("19BJ",option)
        report_df = pd.read_excel(report_location)

        value = xls.sum_column(report_df, "total_chrg_amt")
        cell = db.retrieve_coordinates("19BJ",option)
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
        
        
    except :
        cell = db.retrieve_coordinates("19BJ",option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")
        
        
# needs to be tested
def mmr_by_rate_element(option =''):
    try:

        report_location = db.retrieve_path("19I",option)
        value = pdf.min_rev_by_rate_element(report_location)
        cell = db.retrieve_coordinates("19I",option)
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
        
        
    except :
        cell = db.retrieve_coordinates("19I",option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")
        

def adjustment_occ_report(code='19AU', option =''):
    try:

        report_location = db.retrieve_path(code,option)
        value = pdf.adj_occ_for_AB(report_location)
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
        
        
    except :
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")
        
        
def billing_review(option =''):
    try:
        report_location = db.retrieve_path("19T",option)
        report_df = pd.read_excel(report_location)

        value = xls.sum_column(report_df, "monthly_revenue")
        cell = db.retrieve_coordinates("19T",option)
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
        
        
    except :
        cell = db.retrieve_coordinates("19T",option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")
        
        
def ban_and_clli_error_report(code,option='') :
    try:
        if code == '4F':
            pivot_index = ['ic_cic', 'original_cic']
            sheet = 'BAN Error Report'

        if code == '4H':
            pivot_index = ['ph_number']
            sheet = 'CLLI Error Report'

        
        report_location = db.retrieve_path(code,option)

        if report_location.endswith(".txt") and os.path.getsize(report_location) > 122:
            df = pd.read_csv(report_location, delimiter = "\t",dtype={"fccid": "string", "fg": "string", "ic_cic": "string", "cic_list": "string", "clli": "string",
                                        "ph_number": "string","juris": "string", "traffic_type": "string","clli_co":"string","messages": int,
                                        "minutes": int, "file_name": "string", "file_uid": int, "customer_name": "string", "initial_location": "string","bill_fccid": "string",
                                        "original_cic": "string", "db_parent_cic": "string", "co_parent_cic" : "string", "cic_to_use":"string"})

            try:
                table = pd.pivot_table(df, values='minutes', 
                                index=pivot_index[0], aggfunc=np.sum)

            except:
                table = pd.pivot_table(df, values='minutes', 
                                index=pivot_index[1], aggfunc=np.sum)
            
            table=table.sort_values(by='minutes', ascending=False)
            
            new_file = report_location.replace('.txt', '.xlsx')
            
            with pd.ExcelWriter(new_file) as writer:
                if report_location.find('Peerless_Network') != -1:
                    
                    df1=df.iloc[:df.shape[0]//3,:]
                    df1.to_excel(writer, sheet_name='Sheet_name_1', index=False)

                    df2=df.iloc[df.shape[0]//3:(df.shape[0]//3)*2,:]
                    df2.to_excel(writer, sheet_name='Sheet_name_2', index=False)

                    df3=df.iloc[(df.shape[0]//3)*2:,:]
                    df3.to_excel(writer, sheet_name='Sheet_name_3', index=False)
                    
                else:
                    df1=df.iloc[:,:]
                    df1.to_excel(writer, sheet_name=sheet, index=False)
                    
                table.to_excel (writer,sheet_name="Pivot_Table", index=pivot_index[0], header=True)
                    
            os.remove(report_location)            

            value = xls.sum_column(df1, "messages")
        elif report_location.endswith(".txt") and os.path.getsize(report_location) <= 122:
            df=pd.read_csv(report_location,delimiter="\t")
            new_file = report_location.replace('.txt', '.xlsx')
            df.to_excel(new_file, sheet_name='Report', index=False)
            os.remove(report_location)
            value=0
        elif report_location.endswith(".xlsx"):
            report_df = pd.read_excel(report_location)
            value = xls.sum_column(report_df, "messages")
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")

    except:
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")
        

#ZIP_ERROR should have values as 'ZIP' or 'ERROR'
#Possible values for state in AMBB: 'IN_590G', 'MI_356D', 'OH_509B'
#Possible values for bill type in MTA: 'SW', 'Recip comp'
def zip_and_error_report(company, zip_error, state_billtype = ''):
    try:


        #getting month and year for Q and K drive
        month,year = temp_month,temp_year
        next_month = int(month) + 1
        next_year = year
        bill_type_dic = {'SW':'\\Switched Billing','FA':'\\Facility Billing', 'RC':'\\Recip_Comp'
                            ,'MI_356D':'\\356D_MI','OH_509B':'\\509B_OH' ,'IN_590G':'\\590G_IN','':''}
            
        if next_month in range (1,10):
            next_month = '0' + str(next_month)
        elif next_month == 13:
            next_month = '01'
            next_year = str(int(year)+1)
        else:
            next_month = str(next_month)

        if company != 'Peerless_Network' and company != 'Neutral_Tandem' and company != 'Onvoy':
            #Setting the folder and file paths
            Q_drive_dic = {'BENDTEL':['\\OR_9627', '\\9627_OR_', '22'], 'GCISW':["\\CABS", '\\GCI_SW_', '21'], 
                            'MIEAC':['\\MN_8811', '\\MIEAC_', '01'], 'SELECTRONICS':['\\VT_0069', '\\0069_VT_', '01'],
                            'AMERICANBROADBANDIN_590G':['\\IN_590G', '\\590G_IN_', '25'], 
                            'AMERICANBROADBANDMI_356D':['\\MI_356D', '\\356D_MI_', '25'],
                            'AMERICANBROADBANDOH_509B':['\\OH_509B', '\\509B_OH_' , '20'],
                            'MTASW':['\\CABS', '\\MTA_SW_', '10'], 'MTARC':['\\RECIP_COMP', '\\MTA_RC_', '05']}

            path_zip = "Q:\\" + company + "\\" + year + "_" + month + Q_drive_dic[company+state_billtype][0]
            path_error = "Q:\\" + company + "\\" + next_year + "_" + next_month + Q_drive_dic[company+state_billtype][0]
            standard_name = Q_drive_dic[company+state_billtype][1] + month + Q_drive_dic[company+state_billtype][2] + year[2:4]
            file_name_zip = standard_name + "_5G_Zipped Files.xlsx"
            file_name_error = standard_name + "_5F_Period Error Files.xlsx"
            
            if zip_error.upper() == "ZIP":
                #Extracting list of files from each folder
                files = [f for f in listdir(path_zip) if isfile(join(path_zip, f))]
                
                #Counting total files in the folder
                total_files = len(files)
                value = total_files

                #Creating data frame with the data recovered from folders
                df_zip = pd.DataFrame({'Files':files})
                df_zip.loc[0,'Total Zipped files'] = total_files
                df_zip.loc[0,'Month and Year'] = month + "_" + year

                #Checking if all files were zipped
                total_non_zipped = len(df_zip[['Files']].where(~df_zip['Files'].str.endswith('.zip')).dropna()) #~ as negation
                #GCI
                if company == "GCI":
                    total_non_zipped_ericsson = (len(df_zip[['Files']].where(df_zip['Files'].str.startswith('Ericsson')).dropna())/2) #Counting Ericsson files and taking out err files
                if total_non_zipped != 0:
                    df_zip.loc[0,'Non-Zipped Files'] = total_non_zipped
                    value = "Not all files were zipped."
                    if company == "GCI":
                        value = str(total_files)+"were zipped and "+str(int(total_non_zipped_ericsson))+" Ericsson Files couldnt be zipped"        #it will be pasted in the MPS
            
                #Pasting value in MPS
                cell = db.retrieve_coordinates("5G")
                xls.paste_MPS(value, cell)

                #getting the path where the df will be saved as excel
                if company=="AMERICANBROADBAND":
                    company="American_Broadband"
                path_k='K:\\Client\\' + company + '\\Reports\\' + year + "\\" + month + "_" + year + bill_type_dic[state_billtype]
                
                df_zip.to_excel(path_k+file_name_zip, index = False)

            if zip_error.upper() == "ERROR":
                #Extracting list of files from each folder
                files = [f for f in listdir(path_error) if isfile(join(path_error, f))]

                #Creating data frame with the data recovered from folders
                df_err = pd.DataFrame({'Files':files})
                df_err.loc[0,'Month and Year'] = next_month + "_" + next_year
                if len(files) !=0:

                    #Counting error files in the folder
                    value = len(df_err[['Files']].where(df_err['Files'].str.contains('.err')).dropna())
                    df_err.loc[0,'Total Error files'] = value
                else:
                    value=0
                #Pasting value in MPS
                cell = db.retrieve_coordinates("5F")
                xls.paste_MPS(value, cell)
                #American Broadband fix
                if company=="AMERICANBROADBAND":
                    company="American_Broadband"
                #getting the path where the df will be saved as excel
                path_k='K:\\Client\\' + company + '\\Reports\\' + year + "\\" + month + "_" + year + bill_type_dic[state_billtype] 
                df_err.to_excel(path_k+file_name_error, index = False)
        else:
            company_data_dict= {'Peerless_Network':['PeerlessNetwork','\\Peerless_','05'],
            'Neutral_Tandem':['NTandem','\\Combined_',''],'Onvoy':['Onvoy','\\Onvoy_','01']}
            path_zip='Q:\\'+company_data_dict[company][0]+'\\'+year+'_'+month
            path_error='Q:\\'+company_data_dict[company][0]+'\\'+next_year+'_'+next_month
            file_name_zip=company_data_dict[company][1]+month+company_data_dict[company][2] + year[2:4] + "_5G_Zipped Files.xlsx"
            file_name_error=company_data_dict[company][1]+month+company_data_dict[company][2] + year[2:4] + "_5F_Period Error Files.xlsx"
            folders=[ name for name in os.listdir(path_zip) if os.path.isdir(os.path.join(path_zip, name)) ]
            total_error=0
            total_files=0
            total_non_zipped=0
            if zip_error.upper() == "ZIP":
                zip_total_files=[]
                non_zipped_folders=[]
                non_zipped=False
                for folder in folders:
                    path_folder=path_zip+"\\"+folder
                    
                    #Extracting list of files from each folder
                    files = [f for f in listdir(path_folder) if isfile(join(path_folder, f))]
                    
                    #Checking if all files were zipped
                    if folder != 'badrecords' and folder != 'CABSOUT' and folder != '~PG_2022_03' and len(files) != 0:
                        df_files = pd.DataFrame({'Files':files})
                        total_non_zipped = total_non_zipped+len(df_files[['Files']].where(~df_files['Files'].str.endswith('.zip')).dropna()) #~ as negation
                        if total_non_zipped != 0:
                            non_zipped_folders.append(folder)
                            non_zipped=True
                            total_non_zipped=0
                            continue
                        if non_zipped == False:
                            zip=df_files[['Files']].values.tolist()
                            for i in zip:
                                zip_total_files.append(i[0])
                            error_files=len(df_files[['Files']].where(df_files['Files'].str.endswith('ERR.TXT.zip')).dropna())
                            total_error=total_error+error_files
                            total_files = total_files+(len(files)-error_files)
                    error_files=0
                if non_zipped== True:
                    total_files=''
                    comment="Not all files were zipped or cleaned in folders: "+str(non_zipped_folders)+" please zip them and re run the report"
                elif non_zipped==False:    
                    #Creating data frame with the data recovered from folders
                    df_zip = pd.DataFrame({'Files':zip_total_files})
                    df_zip.loc[0,'Total Zipped files'] = total_files
                    df_zip.loc[0,'Month and Year'] = month + "_" + year
                value = total_files
                if value != '':
                    cells=['D34']
                    aur_files=xls.get_values_from_excel(cells,xls.mps_path)
                    if int(aur_files[0])==value:
                        comment="There were: "+str(total_error)+" error files this month\nAnd the number of files match the AUR"
                        xls.change_color_by_list(xls.mps_path,['D34'], ['blue'])
                    else:
                        comment="There were: "+str(total_error)+" error files this month\nAnd the number of files doesnt match the AUR please check the AUR amount"
                        xls.change_color_by_list(xls.mps_path,['D34'],['red'])
                    path_k='K:\\Client\\' + company + '\\Reports\\' + year + "\\" + month + "_" + year + bill_type_dic[state_billtype]
                    df_zip.to_excel(path_k+file_name_zip, index = False)
                #Pasting value in MPS
                cell = db.retrieve_coordinates("5G")
                xls.paste_MPS(value, cell,comment)
            else:
                if zip_error.upper() == "ERROR":
                    error_total_files=[]
                    error_files=0
                    for folder in folders:
                        path_folder=path_error+"\\"+folder
                        if folder != '~PG_2022_03':
                            files = [f for f in listdir(path_folder) if isfile(join(path_folder, f))]
                            if folder != 'badrecords' and folder != 'CABSOUT' and len(files) != 0 :
                                #Extracting list of files from each folder
                                df_files = pd.DataFrame({'Files':files})
                                error=df_files[['Files']].where(df_files['Files'].str.contains('.err')).dropna().values.tolist()
                                if len(error) != 0:
                                    for i in error:
                                        error_total_files.append(i[0])
                                    #Counting error files in the folder
                                    error_files = error_files + len(df_files[['Files']].where(df_files['Files'].str.contains('.err')).dropna())
                    #Creating data frame with the data recovered from folders
                    df_err = pd.DataFrame({'Files':error_total_files})
                    df_err.loc[0,'Month and Year'] = next_month + "_" + next_year
                    df_err.loc[0,'Total Error files'] = error_files
                    value=error_files
                    comment='N/A'
                    if company == 'Peerless_Network':
                        aur = db.retrieve_path("4A")
                        file_df= pd.read_excel(aur)
                        period_files=file_df[file_df["pd_errors"] > 0]
                        aur_error=period_files['File Name'].count()-1
                        if aur_error==value:
                            comment='Period files total matches the AUR'
                        else:
                            comment='Period files total doesnt match the AUR, period error report: '+str(value)+'AUR: '+str(aur_error)
                    #Pasting value in MPS
                    cell = db.retrieve_coordinates("5F")
                    if company != 'N/A':
                        xls.paste_MPS(value, cell,comment)
                    else:
                        xls.paste_MPS(value, cell)
                    #getting the path where the df will be saved as excel
                    path_k='K:\\Client\\' + company + '\\Reports\\' + year + "\\" + month + "_" + year + bill_type_dic[state_billtype] 
                    df_err.to_excel(path_k+file_name_error, index = False)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
                                
            
    except:
        cell = db.retrieve_coordinates("5G")
        xls.paste_MPS("file is not found", cell)
        cell = db.retrieve_coordinates("5F")
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")
        

def accounting_detailed_report (option =''):
    try:
        report_location = db.retrieve_path("19BA",option)
        report_df = pd.read_excel(report_location)

        value = xls.sum_columns_range(report_df, 'monthly_revenue','occ_revenue')
        cell = db.retrieve_coordinates("19BA",option)
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
        
        
    except :
        cell = db.retrieve_coordinates("19BA",option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")
        

def accounting_report (option =''):
    try:

        report_location = db.retrieve_path("19BB",option)
        report_df = pd.read_excel(report_location)

        value = xls.sum_column(report_df, 'invoice_amount')
        cell = db.retrieve_coordinates("19BB",option)
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
        
        
    except :
        cell = db.retrieve_coordinates("19BB",option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")

def exception_analisys(code='4K',option = ''):
    try:
        report_location = db.retrieve_path(code,option)

        if not report_location.endswith('.xlsx'):
            if report_location.endswith('.txt'):
                report_location=report_location.replace('txt','pdf')
            value=pdf.exception_analisys(report_location)
            txt=report_location.replace('pdf','txt')
            excel=txt.replace('txt','xlsx')
            report=pd.read_csv(txt, sep='\t')
            if value != '0':
                value= report['error_count'].sum()   
            report.to_excel(excel, sheet_name='Report', index=False)
            os.remove(txt)
            cell = db.retrieve_coordinates(code,option)
            xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
        
        
    except :
        cell = db.retrieve_coordinates("4K",option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n") 

def mou_entries(code='4L',option = ''):
    try:
        report_location = db.retrieve_path(code,option)        
        if not report_location.endswith('.xlsx'):
            uid_values = xls.get_values_from_excel(list_cells = ['C12','C13'],report_location= xls.mps_path)
            value, comment = pdf.mou_entries(report_location, uid_values)
            cell = db.retrieve_coordinates("4L",option)
            xls.paste_MPS(value, cell, comment)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
    except :
        cell = db.retrieve_coordinates("4L",option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")

def adjs_not_posted(code='9F',option=''):
    try:
        report_location = db.retrieve_path(code,option)
        
        value, comment = pdf.adjs_not_posted(report_location)
        cell = db.retrieve_coordinates("9F",option)
        xls.paste_MPS(value, cell, comment)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
    except :
        cell = db.retrieve_coordinates("9F",option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")

def bill_completion(code='12A',option=''):
    try:
        report_location = db.retrieve_path(code,option)     
        value, comment = pdf.bill_completion(report_location)
        cell = db.retrieve_coordinates("12A",option)
        xls.paste_MPS(value, cell, comment)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
    except :
        cell = db.retrieve_coordinates("12A",option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")

def active_bans(code='12F',option=''):
    try:
        report_location = db.retrieve_path(code,option)
        report_df = pd.read_excel(report_location)

        count_active = xls.count_rows(report_df,"active","Y")
        report_location_swbt = db.retrieve_path('12C',option)
        report_df_swbt = pd.read_excel(report_location_swbt)
        count_swbt = xls.count_rows(report_df_swbt,"ban")
        if count_active == count_swbt :
            value=count_active
            comment="The active BANs matches the SWBT"
        else:
            value="The active BANs doesnt match the SWBT"
            comment=0
        cell = db.retrieve_coordinates("12F",option)
        xls.paste_MPS(value,cell,comment)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
    except :
        cell = db.retrieve_coordinates("12F",option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")

def mmr_bill_date(code='19K',option=''):
    try:
        report_location = db.retrieve_path(code,option)
        report_df = pd.read_excel(report_location)

        value = xls.sum_column(report_df,"total_rev")
        cell = db.retrieve_coordinates("19K",option)
        xls.paste_MPS(value, cell, comment=0)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
    except :
        cell = db.retrieve_coordinates("19K",option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")

def gl_extract(code='19Z', option=''):
    try:
        report_location = db.retrieve_path(code,option)
        report_df = pd.read_excel(report_location)

        value = xls.sum_column(report_df,"revenue")
        mou=xls.sum_column(report_df,"mou")
        cell = db.retrieve_coordinates("19Z",option)
        xls.paste_MPS(value, cell, comment="MOU: "+str(mou))
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
    except :
        cell = db.retrieve_coordinates("19Z",option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")

def soc_factor_change(code,option=''):
    try:
        report_location = db.retrieve_path(code,option)
        value, comment = pdf.soc_factor_change(code,report_location)
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS(value, cell, comment)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
    except :
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")

def number_of_invoices(code='21B',option=''):
    try:
        report_location = db.retrieve_path(code,option)
        report_df = pd.read_excel(report_location)
        try:
            total=report_df['total_amount_due'].sum()
            value=len(report_df['total_amount_due'])
        except:
            total=report_df['total_amount'].sum()
            value=len(report_df['total_amount'])                                    
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS(value, cell)
        cell = db.retrieve_coordinates('21C',option)
        xls.paste_MPS(total, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
    except :
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")
        
def aurs_bill_date(code,option=''):
    try:
        report_location = db.retrieve_path(code,option)
        if report_location.endswith('.txt') == True :
            report_df = pd.read_csv(report_location, sep="\t",header=0)
            new_report=report_location.replace('.txt','.xlsx')
            report_df.to_excel(new_report,index=False)
            value='X'
            cell = db.retrieve_coordinates(code,option)
            xls.paste_MPS(value, cell)
            os.remove(report_location)
            print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
            db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
        elif report_location.endswith('.xlsx') == True :
            value='X'
            cell = db.retrieve_coordinates(code,option)
            xls.paste_MPS(value, cell)
            print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
            db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
    except :
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")

def occ_billed(code = '19BK',option = ''):
    try:
        report_location = db.retrieve_path(code,option)
        report_df = pd.read_excel(report_location)
        value=report_df['amount_billed'].sum()
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
    except :
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")

def rev_by_clli(code='19BC',option=''):
    try:
        report_location = db.retrieve_path(code,option)
        report_df = pd.read_excel(report_location)
        value=report_df['monthly_revenue'].sum()
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is done\n")
    except :
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")
        db.log_file.write(datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ' | ' + sys._getframe().f_code.co_name + " is not found\n")

def soc_jurisdiction(code='19AW',option=''):
    try:
        report_location = db.retrieve_path(code,option)
        report_df = pd.read_excel(report_location)
        value = xls.sum_columns_range(report_df, 'iser','local')
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
    except :
        cell = db.retrieve_coordinates("19AW",option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")

def rev_analysis7(code='19R',option=''):
    try:
        report_location = db.retrieve_path(code,option)
        report_df = pd.read_excel(report_location)
        value=report_df['monthly_revenue'].sum()
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
    except :
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")

def Switch_Terminating_Intrastate_Rev_Report_for_NECA(code = '19AJ', option = ''):
    
    try:
        report_location = db.retrieve_path(code,option)
        report_df = pd.read_excel(report_location, converters={'bill_fccid':str}).dropna(subset=['bill_fccid'])
        df1 = xls.add_filter(report_df, ['direction'], ['T'])
        df2 = xls.add_filter(report_df,['rateeleid_emrttid','direction'], ['CCLT','O'] )
        xls.add_sheet(report_location, ['msg_mou_rev_' + str(temp_month) + str(temp_year),'Intrastate Terminating Rev','direction O'], [report_df, df1,df2], 3)
        value = 'x'
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS(value, cell)
        print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
    except :
        cell = db.retrieve_coordinates(code,option)
        xls.paste_MPS("file is not found", cell)
        print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")

def pre_deletions (code= '15A' ,option=''):
    try:
        report_location = db.retrieve_path("15A",option)
        report_df = pd.read_excel(report_location)
        to_delete=report_df[(report_df["prev_balance"] == 0) & (report_df["total_current_chgs"] != 0) & (report_df["total_balance_due"] < 5) ]
        if to_delete.empty:
            value = 'Total amount: 0 Total BANs: 0'
            cell = db.retrieve_coordinates("15",option)
            xls.paste_MPS(value, cell)
            value = 'N/A'
            cell = db.retrieve_coordinates("15A",option)
            xls.paste_MPS(value, cell)
            print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        else:
            to_delete=to_delete[["ban",'prev_balance','total_current_chgs','total_balance_due']]
            total=to_delete["total_balance_due"].sum()
            total_bans=to_delete["ban"].count()
            bans=list(to_delete["ban"])
            with pd.ExcelWriter(report_location, engine='openpyxl', mode='a') as writer:  
                to_delete.to_excel(writer, sheet_name='To Delete',index=False)
            comment="To delete: \n" + to_delete.to_string(index=False)
            value="Total amount: " +str(total)+" Total BANs: "+ str(total_bans)#15 value
            value2=str(bans)#15A value
            cell = db.retrieve_coordinates("15",option)
            xls.paste_MPS(value, cell,comment)
            cell = db.retrieve_coordinates("15A",option)
            xls.paste_MPS(value2, cell)
            print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
    except:
            cell = db.retrieve_coordinates(code,option)
            xls.paste_MPS("file is not found", cell)
            print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")

def pos_deletions(code= '15F' ,option=''):
    try:
        not_found=[]
        bans=load_workbook(xls.mps_path).active['D119'].value
        bans=bans.replace('[','')
        bans=bans.replace(']','')
        bans=bans.replace('\'','')
        bans=list(bans.split(','))
        if bans != 'N/A':
            report_location=db.retrieve_path("15F",option)
            report_df = pd.read_excel(report_location)
            zero=report_df[(report_df["total_balance_due"] == 0)]
            pos_bans=list(zero["ban"])
            for ban in bans:
                ban=ban.replace(" ","")
                if ban not in pos_bans:
                    not_found.append(ban)
            if not_found:
                value="Please check these BANs: " + str(not_found) #Value
                cell = db.retrieve_coordinates("15F",option)
                xls.paste_MPS(value, cell)
                print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")

            else:
                value=len(bans) #Value
                cell = db.retrieve_coordinates("15F",option)
                xls.paste_MPS(value, cell)
                print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
        else:
            value="N/A"
            cell = db.retrieve_coordinates("15F",option)
            xls.paste_MPS(value, cell)
            print("\033[1;32;40m" +sys._getframe().f_code.co_name + " is done" + "\033[0;37;40m")
    except:
            cell = db.retrieve_coordinates(code,option)
            xls.paste_MPS("file is not found", cell)
            print("\033[1;31;40m" + sys._getframe().f_code.co_name + " is not found" + "\033[0;37;40m")

def macros(macro):
   
    xl = win32com.client.Dispatch('Excel.Application')
    wb = xl.Workbooks.Open(os.path.abspath(macro_path+'\\Peerless_Network_Macros.xlsm'))
    xl.Application.Run('Peerless_Network_Macros.xlsm'+"!Module1."+macro)
    wb.Save
    wb.Close()
    del xl