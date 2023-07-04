from openpyxl import load_workbook
import numpy as np
import pandas as pd

def format_tbl(writer, sheet_name, df):
    outcols = df.columns
    if len(outcols) > 25:
        raise ValueError('table width out of range for current logic')
    tbl_hdr = [{'header':c} for c in outcols]
    bottom_num = len(df)+1
    right_letter = chr(65-1+len(outcols))
    tbl_corner = right_letter + str(bottom_num)

    worksheet = writer.sheets[sheet_name]
    worksheet.add_table('A1:' + tbl_corner,  {'columns':tbl_hdr})

def fill_dashboard(dashboard,report_mps,client,bill_type,month,year,mps2=""): # 1=GCI, BENDTEL, MTA || 2=ABB  
    wb = load_workbook(report_mps,data_only=True)
    sh = wb["MPS"]
    dash_workbook = load_workbook(dashboard)
    dash_sheet = dash_workbook[month+year]
    cells={'PEERLESSSW':['184'],'NEUTRAL TANDEMSW':['15'],'NEUTRAL TANDEMFA':['14'],'ONVOYSW':['111'],'MTASW':['12'],'MTARC':['11']
    ,'GCISW':['9'],'GCIFA':['8'],'MIEACSW':['13'],'ONVOYSW':['111'],'AMBBMI':['5'],'AMBBOH':['3']
    ,'AMBBIN':['4'],'BENDTELSW':['6']}
            
    if client+bill_type=="MTASW":
        wb2 = load_workbook(mps2,data_only=True)
        sh2 = wb2["MPS"]
        alldt=[sh['D18'].value,sh2['D18'].value]
        prod1=[sh['D30'].value,sh2['D30'].value]
        setup=[sh['D85'].value,sh2['D85'].value]
        bdone=[sh['D230'].value,sh2['D230'].value]
        sroff=[sh['H230'].value,sh2['H230'].value]
        datef=[sh['D238'].value,sh2['D238'].value]
        all=np.max(alldt)
        prod=np.max(prod1)
        set=np.max(setup)
        bdon=np.max(bdone)
        srof=np.max(sroff)
        date=np.max(datef)
        if sh['D230'].value >= sh2['D230'].value:
            bdoneby= sh['G230'].value
        else:
            bdoneby= sh2['G230'].value
        if sh['H230'].value >= sh2['H230'].value:
            sroffby= sh['I230'].value
        else:
            sroffby= sh2['I230'].value
        total_bills=str(int(sh['D228'].value)+int(sh2['D228'].value))
        dash_sheet['J'+cells[client+bill_type][0]] = all
        dash_sheet['K'+cells[client+bill_type][0]] = prod
        dash_sheet['L'+cells[client+bill_type][0]] = set
        dash_sheet['M'+cells[client+bill_type][0]] = bdon
        dash_sheet['N'+cells[client+bill_type][0]] = srof
        dash_sheet['O'+cells[client+bill_type][0]] = date
        dash_sheet['P'+cells[client+bill_type][0]] = total_bills
        dash_sheet['Q'+cells[client+bill_type][0]] = sroffby
        dash_sheet['R'+cells[client+bill_type][0]] = bdoneby
        dash_sheet['AE'+cells[client+bill_type][0]] = sh['D34'].value
        dash_sheet['AF'+cells[client+bill_type][0]] = sh['D32'].value


        df = pd.DataFrame({'Client':[client],
                           'All Data Recorded':[str(all)],
                           'Prod. Day One':[str(prod)],
                           'Setup & Produce Done':[str(set)],
                           'Billing Done':[str(bdon)],
                           'Sr. Sign-Off':[str(srof)],
                           'Date Fulfilled':[str(date)],
                           'No. of Bills': [total_bills],
                           'Sr. Sign Off By':[sroffby],
                           'Bill Done By':[bdoneby],
                           'Total Files':[sh['D34'].value],
                           'Total Records':[sh['D32'].value]})
        update_path=dashboard.replace(".xlsx"," - Update.xlsx")  
        update_path=update_path.replace("Production Dashboard","Production Dashboard\\Update teams")
        with pd.ExcelWriter(update_path, mode='w',date_format='YYYY-MM-DD',datetime_format='YYYY-MM-DD HH:MM:SS',engine='xlsxwriter') as writer:
                       
            sheet_name='Sheet1'
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            format_tbl(writer, sheet_name, df)
        #close the workbook after reading
        wb2.close()
        wb2 = load_workbook(mps2)
        sh2 = wb2["MPS"]

        sh2['D236']='X'
        wb2.save(mps2)

        #close the workbook after reading
        wb2.close()





    if client+bill_type!="MTASW":
        #pasting the value in a specific cell
        dash_sheet['J'+cells[client+bill_type][0]] = sh['D18'].value
        dash_sheet['K'+cells[client+bill_type][0]] = sh['D30'].value
        dash_sheet['L'+cells[client+bill_type][0]] = sh['D85'].value
        dash_sheet['M'+cells[client+bill_type][0]] = sh['D230'].value
        dash_sheet['N'+cells[client+bill_type][0]] = sh['H230'].value
        dash_sheet['O'+cells[client+bill_type][0]] = sh['D238'].value
        dash_sheet['P'+cells[client+bill_type][0]] = sh['D228'].value
        dash_sheet['Q'+cells[client+bill_type][0]] = sh['I230'].value
        dash_sheet['R'+cells[client+bill_type][0]] = sh['G230'].value
        dash_sheet['AE'+cells[client+bill_type][0]] = sh['D34'].value
        dash_sheet['AF'+cells[client+bill_type][0]] = sh['D32'].value
        if bill_type=='FA':
            dash_sheet['AE'+cells[client+bill_type][0]] = 'N/A'
            dash_sheet['AF'+cells[client+bill_type][0]] = 'N/A'
        if bill_type=='FA':
            tfiles = 'N/A'
            trecords = 'N/A'
        else:
            tfiles = sh['D34'].value
            trecords = sh['D32'].value

        df = pd.DataFrame({'Client':[client],
                           'All Data Recorded':[str(sh['D18'].value)],
                           'Prod. Day One':[str(sh['D30'].value)],
                           'Setup & Produce Done':[str(sh['D85'].value)],
                           'Billing Done':[str(sh['D230'].value)],
                           'Sr. Sign-Off':[str(sh['H230'].value)],
                           'Date Fulfilled':[str(sh['D238'].value)],
                           'No. of Bills': [sh['D228'].value],
                            'Sr. Sign Off By':[sh['I230'].value],
                           'Bill Done By':[sh['G230'].value],
                           'Total Files':[tfiles],
                           'Total Records':[trecords]})
        update_path=dashboard.replace(".xlsx"," - Update.xlsx")  
        update_path=update_path.replace("Production Dashboard","Production Dashboard\\Update teams")
        with pd.ExcelWriter(update_path, mode='w',date_format='YYYY-MM-DD',datetime_format='YYYY-MM-DD HH:MM:SS',engine='xlsxwriter') as writer:
            sheet_name='Sheet1'
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            format_tbl(writer, sheet_name, df)
    # if client=='PEERLESS' or client=='GCI' or client=='MTA':
    #     dash_sheet['O'+cells[client+bill_type][0]]='' 
    #     print("Release to print section must be filled once billing is approved by client")
    
        

    #close the workbook after reading
    wb.close()
    wb = load_workbook(report_mps)
    sh = wb["MPS"]

    sh['D236']='X'
    wb.save(report_mps)

    #close the workbook after reading
    wb.close()


    #saving the spreadsheet
    dash_workbook.save(dashboard)

    #close the workbook after reading
    dash_workbook.close()

    print("Dashboard done, remember to have the COMPLETE FULLFILMENT field filled on the MPS (D238)")

# dashboard="K:\\Client\\Service_bureau\\Audit\\Production Dashboard\\Production_Dashboard_Data_2022.xlsx"
# mps=r"K:\Client\MIEAC\Reports\2022\11_2022\MPS_MN_8811_11_22.xlsx"
# client='MIEAC'
# bill_type='SW'
# month="Nov"
# year="22"
def call(dashboard,mps,client,bill_type,month,year,mps2=""):
    fill_dashboard(dashboard,mps,client,bill_type,month,year,mps2)
