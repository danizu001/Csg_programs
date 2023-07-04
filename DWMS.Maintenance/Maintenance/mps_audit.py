from genericpath import isfile
from openpyxl import load_workbook
import openpyxl.utils.cell
from zipfile import ZipFile
import os
import shutil



def audit_mps(company,mps,log):
    wb = load_workbook(mps,data_only=True) 
    ws = wb['MPS']
    empty=[]
    if ws["C12"].value == None:
        empty.append('C12')
    if ws["C13"].value == None:
        empty.append('C13')
    for row in ws["C17":"I268"]:
        if ws.row_dimensions[row[0].row].hidden == False:
            for cell in row:
                if company != "Neutral_Tandem_SW":
                    if openpyxl.utils.cell.get_column_letter(cell.column) != "E" and openpyxl.utils.cell.get_column_letter(cell.column) != "F":
                        if cell.value == None:
                            empty.append(cell.coordinate)
                else:
                    if cell.value == None:
                        empty.append(cell.coordinate)
    if empty:
        print("Billing done by: "+str(ws["C6"].value))
        log.append("Billing done by: "+str(ws["C6"].value)+"\n")
        print("The Following Cells are not filled: ")
        log.append("The Following Cells are not filled: \n")
        print(empty)
        log.append(str(empty)+"\n")
    else:
        print("Everything filled")
        log.append("Everything filled\n")
def call(company,btype,year):
    btypes={'SW':['Switched Billing'],'FA':['Facility Billing'], 'RC':['Recip_Comp'],'IN':['590G_IN'],'OH':['509B_OH'],'MI':['356D_MI'],'':['']}
    months=["01","02","03","04","05","06","07","08","09","10","11","12"]
    year_path="K:\\Client\\"+company+"\\Reports\\"+year+"\\Test\\"
    file1 = open(year_path+"MPS Audit Report "+btype+".txt","w")
    file1.write(company+"\t"+year+"\t"+"--------AUDIT REPORT------------\n")
    file1.close() #to change file access modes
    for month in months:
        files=[]
        mps_path=year_path+month+"_"+year+"\\"+btypes[btype][0]
        log=[]
        print("Auditing month: "+month)
        log.append("Auditing month: "+month+"\n")
        try:
            archives=os.listdir(mps_path)
            for file in archives:
                if file.endswith('xlsx') or file.endswith('pdf'):
                    files.append(file)
            if len(files) != 0 and len(files) > 3:
                for file in files:
                    if file.startswith("MPS"):
                        mps=os.path.join(mps_path, file)
                try:        
                    audit_mps(company,mps,log)
                except:
                    print("There is an error with the file, might be corrupted, please check")
                    log.append("There is an error with the file, might be corrupted, please check\n")

            elif len(files) == 0:
                print("Month: " +month+" has no MPS")
                log.append("Month: " +month+" has no MPS\n")

            else:    
                print("Month: " +month+" has MPS but billing has not started")
                log.append("Month: " +month+" has MPS but billing has not started\n")

        except:
                if company != "American_Broadband":
                    zipped_path=year_path+month+"_"+year+"-"+year+"-"+month+".zip"
                else:
                    zipped_path=year_path+"Reports-"+year+"-"+month+".zip"
                with ZipFile(zipped_path, 'r') as zipObject:
                    listOfFileNames = zipObject.namelist()
                    for fileName in listOfFileNames:
                        if fileName.startswith(btypes[btype][0]+'/MPS') or fileName.startswith("MPS") :
                        # Extract all the contents of zip file in current directory
                            zipObject.extract(fileName, year_path+month+"_"+year+"\\")
                            print(month+' is zipped and file has been extracted')    
                            log.append(month+' is zipped and file has been extracted\n')
            
                for file in os.listdir(mps_path):
                    if file.startswith("MPS"):
                        mps=os.path.join(mps_path, file)
                    try:        
                        audit_mps(company,mps,log)
                    except:
                        print("There is an error with the file, might be corrupted, please check")
                        log.append("There is an error with the file, might be corrupted, please check\n")
                #shutil.rmtree(mps_path)
        file1 = open(year_path+"MPS Audit Report "+btype+".txt","a")
        log.append("---------------------------\n")
        file1.writelines(log)
        file1.close() #to change file access modes

