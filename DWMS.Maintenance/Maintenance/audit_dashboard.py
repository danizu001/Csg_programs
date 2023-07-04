from genericpath import isfile
from openpyxl import load_workbook
import openpyxl.utils.cell
from zipfile import ZipFile
import os
import shutil

def audit_dash(dashboard,log,year):
    wb = load_workbook(dashboard,data_only=True) 
    yr=str(year)[-2:]
    months=['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    for month in months:
        ws = wb[month+yr]
        empty=[]
        print("Working on month: "+month)
        log.append("Working on month: "+month+"\n")
        for row in ws["J3":"O243"]:
            if ws.row_dimensions[row[0].row].hidden == False:
                for cell in row:
                    getline=str(cell.coordinate)[-1]
                    if cell.value == None and getline != '6' :
                        empty.append(cell.coordinate)
        for row in ws["P3":"R15"]:
            if ws.row_dimensions[row[0].row].hidden == False:
                for cell in row:
                    getline=str(cell.coordinate)[-1]
                    if cell.value == None and getline != '6' and getline != '8' and getline != '0': #10
                        empty.append(cell.coordinate)
        for row in ws["P111":"R111"]:
            if ws.row_dimensions[row[0].row].hidden == False:
                for cell in row:
                    getline=str(cell.coordinate)[-1]
                    if cell.value == None :
                        empty.append(cell.coordinate)
        for row in ws["P184":"R184"]:
            if ws.row_dimensions[row[0].row].hidden == False:
                for cell in row:
                    getline=str(cell.coordinate)[-1]
                    if cell.value == None :
                        empty.append(cell.coordinate)
        for row in ws["AE3":"AF15"]:
            if ws.row_dimensions[row[0].row].hidden == False:
                for cell in row:
                    getline=str(cell.coordinate)[-1]
                    if cell.value == None and getline != '6' and getline != '0' :
                        empty.append(cell.coordinate)

        for row in ws["AE111":"AF111"]:
            if ws.row_dimensions[row[0].row].hidden == False:
                for cell in row:
                    getline=str(cell.coordinate)[-1]
                    if cell.value == None :
                        empty.append(cell.coordinate)
        for row in ws["AE184":"AF184"]:
            if ws.row_dimensions[row[0].row].hidden == False:
                for cell in row:
                    getline=str(cell.coordinate)[-1]
                    if cell.value == None :
                        empty.append(cell.coordinate)
        if empty:
            print("The Following Cells are not filled or the month has not started yet: ")
            log.append("The Following Cells are not filled or the month has not started yet: \n")
            print(empty)
            empty.sort()
            log.append(str(empty)+"\n")
        else:
            print("Everything filled")
            log.append("Everything filled\n")


def call(year):
    dash_path='K:\\Client\\Service_bureau\\Audit\\Production Dashboard\\'
    dashboard='K:\\Client\\Service_bureau\\Audit\\Production Dashboard\\Production_Dashboard_Data_'+str(year)+'.xlsx'
    file1 = open(dash_path+"Dashboard Audit Report "+".txt","w")
    file1.write("\t"+year+"\t"+"--------AUDIT REPORT------------\n")
    file1.close() #to change file access modes
    log=[]
    audit_dash(dashboard,log,year)
    file1 = open(dash_path+"Dashboard Audit Report "+".txt","a")
    log.append("---------------------------\n")
    file1.writelines(log)
    file1.close() #to change file access modes



