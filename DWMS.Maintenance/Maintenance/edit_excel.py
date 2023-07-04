from openpyxl import load_workbook
import pandas as pd
column_month=['B','C','D','E','F','G','H','I','J','K','L','M']
dict_comp={'American_Broadband':'3','BendTel':'4','GCI':'5','MIEAC':'6','MTA':'7','Neutral_Tandem':'8','ONVOY':'9','Peerless_Network':'10'}
def call(folder,mps_list,company, month, name, revenue, revenue_name):
    volume_wb = load_workbook(folder+name)
    Revenue_sh = volume_wb['Revenue Trends']
    Message_sh = volume_wb['Messages Trends']
    sum_msg=[]
    sum_rev=[]
    if company=='GCI' or company=='Neutral_Tandem' :
        for i in mps_list:
            wb = load_workbook(i[0]+i[1],data_only=True)
            sh = wb["MPS"]
            if 'Facility' in i[0]:
                pass
                # sum_rev.append(sh['D183'].value)
            else:
                sum_msg.append(sh['D32'].value)
                # sum_rev.append(sh['D162'].value)
            wb.close()
    elif company=='MTA':
        for i in mps_list:
            wb = load_workbook(i[0]+i[1],data_only=True)
            sh = wb["MPS"]
            if 'Facility' in i[0]:
                sum_rev.append(sum_excel(i[0]+revenue[1][0]+revenue[1][1]+revenue_name))
                # sum_rev.append(sh['D183'].value)
            elif 'Recip' in i[0]:
                sum_rev.append(sum_excel(i[0]+revenue[2][0]+revenue[2][1]+revenue_name))
                sum_msg.append(sh['D32'].value)
                # sum_rev.append(sh['D162'])
            else:
                sum_msg.append(sh['D32'].value)
                # sum_rev.append(sh['D162'].value)
            wb.close()
    elif company=='Peerless_Network':
        wb = load_workbook(mps_list[0][0]+mps_list[0][1],data_only=True)
        sh = wb["MPS"]
        revenue_name=revenue_name.replace('Revenue Analysis','Revenue Analysis 7')
        sum_msg.append(sh['D32'].value)
        sum_rev.append(sum_excel(mps_list[0][0]+revenue[0][0]+revenue[0][1]+revenue_name))
            # sum_rev.append(sh['D162'].value)
        wb.close()
    else:
        for i in range(len(mps_list)):
            wb = load_workbook(mps_list[i][0]+mps_list[i][1],data_only=True)
            sh = wb["MPS"]
            sum_msg.append(sh['D32'].value)
            if company=='ONVOY' or company=='MIEAC':
                sum_rev.append(sum_excel(mps_list[i][0]+revenue[i][0]+revenue[i][1]+revenue_name))
            else:
                sum_rev.append(sh['D219'].value)
            wb.close()
    Revenue_sh[column_month[int(month)-1]+dict_comp[company]]=sum(sum_rev)
    Message_sh[column_month[int(month)-1]+dict_comp[company]]=sum(sum_msg)
    volume_wb.save(folder+name)
    volume_wb.close()
    
def sum_excel(name):
    df=pd.read_excel(name)
    revenue_sum=sum(df['monthly_revenue'])+sum(df['occ_revenue'])
    return revenue_sum        
        
