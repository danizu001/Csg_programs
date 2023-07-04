import pandas as pd
import os
from openpyxl import load_workbook
directory="K:\\Client\\Peerless_Network\\Reports\\2023\\04_2023\\"
def pre_threshold(directory):
    for file in os.listdir(directory):
        if "14A" in file:
            df=pd.read_excel(directory+file)
            dfbans = df[(df["usage_chrg"].between(.01, 24.99)) & (df["threshold_billing"] == "N") ]
            if not dfbans.empty:
                bans=list(dfbans["ban"])
                print("Please mark the following bans prior to threshold: ")
                print(bans)
            else:
                df= df[(df["usage_chrg"].between(.01, 24.99))]
                bans=list(df['ban'])
                totalbans=len(bans)
                totalamount=round(sum(df['usage_chrg']),3)
                write_mps(directory,totalbans,totalamount)
                print("All Bans are marked, procceed with billing, total Bans: ",len(bans), " total amount :", round(sum(df['usage_chrg']),3))

def pos_threshold(directory):
    for file in os.listdir(directory):
        if "14C" in file:
            df=pd.read_excel(directory+file)
            totalbans=df['ban'].count()
            for file in os.listdir(directory):
                if "MPS" in file:
                    wb = load_workbook(directory+file)
                    sh = wb["MPS"]
                    totalmps=sh['D115'].value
                    wb.close()
            if totalbans == int(totalmps):
                write_mps(directory,totalbans,prethreshold='no')
                print("All Bans were thresheld correctly")
            else:
                for file in os.listdir(directory):
                    if "14A" in file:
                        pre=pd.read_excel(directory+file)
                        dfbans = pre[(pre["usage_chrg"].between(.01, 24.99))]
                        prebans=list(dfbans["ban"])
                        posbans=list(df["ban"])
                        missing=[]
                        if len(prebans) > len(posbans):
                            for ban in prebans:
                                if ban not in posbans:
                                    missing.append(ban)
                        else:
                            for ban in posbans:
                                if ban not in prebans:
                                    missing.append(ban)
                        print("Please check the following BANs: ")
                        print(missing)


def write_mps(directory,totalbans=0,totalamount=0,prethreshold="yes"):
    for file in os.listdir(directory):
        if "MPS" in file:

            wb = load_workbook(directory+file)
            sh = wb["MPS"]
            if prethreshold=='yes':
                sh['D115']=str(totalbans)
                sh['D114']=str(totalamount)
            else:
                sh['D117']=str(totalbans)
            wb.save(directory+file)

            #close the workbook after reading
            wb.close()

