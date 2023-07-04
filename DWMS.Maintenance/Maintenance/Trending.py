from openpyxl import load_workbook
import pdfplumber
import re
from win32com import client

def data_mps(report_mmr,report_mps,client): # 1=GCI, BENDTEL, MTA || 2=ABB  
    data=[]
    wb = load_workbook(report_mps, data_only=True)
    sh = wb["MPS"]
    data.append(sh['D35'].value)
    data.append(sh['D34'].value)
    data.append(str(sh['C12'].value)+'-'+str(sh['C13'].value))
    if(client==1):
        data.append(round(float(sh['D100'].value)+float(sh['D101'].value),2))
    elif(client==2):
        data.append(sh['D102'].value)
    data2=extract_data_pdf(report_mmr)
    return data,data2

def extract_data_pdf(report_mmr):
    data=[]
    with pdfplumber.open(report_mmr) as pdf:
        file = pdf.pages[-1].extract_text()  # reading the text of the last page of the report
        text = file[file.find("Grand Totals") : -1]  # getting the text part of the last table "Grand Totals"
        msg = text[text.find("Total Messages") : -1]  # getting the text part of the revenue
        msg_split = msg.split()  # spliting the numbers into list
        mou = text[text.find("Total Minutes") : -1]  # getting the text part of the revenue
        mou_split = mou.split()  # spliting the numbers into list
        revenue = text[text.find("Total Revenue") : -1]  # getting the text part of the revenue
        revenue_split = revenue.split()  # spliting the numbers into list
    data=[msg_split[9],mou_split[9],revenue_split[9],mou_split[4],revenue_split[4],mou_split[7],revenue_split[7],mou_split[8],revenue_split[8]]
    return data

def paste_values_trending_and_mps(report_BC,line,data,data2,sheet_name,client):# 1=ABB,Bendtel || 2=GCI,MTA
    wb = load_workbook(report_BC)
    sh = wb[sheet_name]
    sh['B'+str(line)]=float(re.sub("[^\d\.]", "", str(data[0]).replace("$", "")))
    sh['C'+str(line)]=data[1]
    sh['D'+str(line)]=data[2]
    sh['F'+str(line)]=float(re.sub("[^\d\.]", "", data2[0].replace("$", "")))
    sh['U'+str(line)]=float(re.sub("[^\d\.]", "", data2[5].replace("$", "")))
    sh['V'+str(line)]=float(re.sub("[^\d\.]", "", data2[6].replace("$", "")))
    if client == 1:
        sh['I'+str(line)]=float(re.sub("[^\d\.]", "", data2[1].replace("$", "")))
        sh['J'+str(line)]=float(re.sub("[^\d\.]", "", data2[7].replace("$", "")))
        sh['L'+str(line)]=float(re.sub("[^\d\.]", "", data2[2].replace("$", "")))
        sh['M'+str(line)]=data[3]
        sh['R'+str(line)]=float(re.sub("[^\d\.]", "", data2[3].replace("$", "")))
        sh['S'+str(line)]=float(re.sub("[^\d\.]", "", data2[4].replace("$", "")))
        sh['X'+str(line)]=float(re.sub("[^\d\.]", "", data2[7].replace("$", "")))
        sh['Y'+str(line)]=float(re.sub("[^\d\.]", "", data2[8].replace("$", "")))
        adj_revenue=(sh['L'+str(line)].value-sh['M'+str(line)].value)
    elif client == 2:
        sh['H'+str(line)]=float(re.sub("[^\d\.]", "", data2[1].replace("$", "")))
        sh['I'+str(line)]=float(re.sub("[^\d\.]", "", data2[2].replace("$", "")))
        sh['K'+str(line)]=data[3]
        sh['P'+str(line)]=float(re.sub("[^\d\.]", "", data2[3].replace("$", "")))
        sh['Q'+str(line)]=float(re.sub("[^\d\.]", "", data2[4].replace("$", "")))
        sh['Z'+str(line)]=float(re.sub("[^\d\.]", "", data2[7].replace("$", "")))
        sh['AA'+str(line)]=float(re.sub("[^\d\.]", "", data2[8].replace("$", "")))
        adj_revenue=(sh['I'+str(line)].value+sh['J'+str(line)].value-sh['K'+str(line)].value)
    wb.save(report_BC)
    wb.close()
    return adj_revenue
    
def Prepare_data(report_mps,reference,date,prelim='',prelim2='',sheet_name='',client=0):# 1=GCI MTA, 2=ABB, 3 BENDTEL
    split=report_mps.split('\\')
    if client==1:
        path_mmr=split[:7]
        path_mmr.append(prelim+split[5][:2]+date+split[5][5:]+"_19G_Message MOU Revenue by BAN Report.pdf")
        report_mmr='\\'.join(path_mmr)
        path_BC=split[:5]
        path_BC.append(split[4]+prelim2)
        report_BC='\\'.join(path_BC)
        data,data2=data_mps(report_mmr, report_mps,1)
        number=int(split[5][:2]) #take a number of the month for know where we need to write in the excel
        adj_revenue = paste_values_trending_and_mps(report_BC, reference+number, data, data2, "Roll Up", 2)
    elif client==2:
        path_mmr=split[:7]
        path_mmr.append(split[6]+"_"+split[5][:2]+date+split[5][5:]+"_19G_Message MOU Revenue by BAN Report.pdf")
        report_mmr='\\'.join(path_mmr)
        path_BC=split[:5]
        path_BC.append('Trending_American Broadband_'+split[4]+'.xlsx')
        report_BC='\\'.join(path_BC)
        data,data2=data_mps(report_mmr, report_mps,2)
        number=int(split[5][:2]) #take a number of the month for know where we need to write in the excel
        adj_revenue = paste_values_trending_and_mps(report_BC, reference+number, data, data2, "Trending-"+split[4], 1)
    elif client==3:
        path_mmr=split[:6]
        path_mmr.append(prelim+split[5][:2]+date+split[5][5:]+"_19G_Message MOU Revenue by BAN Report.pdf")
        report_mmr='\\'.join(path_mmr)
        path_BC=split[:5]
        path_BC.append(prelim2+split[4]+'.xlsx')
        report_BC='\\'.join(path_BC)
        data,data2=data_mps(report_mmr, report_mps,1)
        number=int(split[5][:2]) #take a number of the month for know where we need to write in the excel
        adj_revenue = paste_values_trending_and_mps(report_BC, reference+number, data, data2, sheet_name, 1)
    paste_mps(report_mps,adj_revenue)
    save_as_pdf(report_BC,report_mmr) 

def paste_mps(report_mps,adj_revenue):
    #opening the MPS spreadsheet and selecting the main sheet
    MPS_workbook = load_workbook(report_mps)
    MPS_sheet = MPS_workbook.active
    
    #pasting the value in a specific cell
    MPS_sheet['D219'] = adj_revenue

    #saving the spreadsheet
    MPS_workbook.save(report_mps)

    #close the workbook after reading
    MPS_workbook.close()


def save_as_pdf(report_bc,mmr_report):
    
    # Open Microsoft Excel
    excel = client.Dispatch("Excel.Application")
    
    # Read Excel File
    sheets = excel.Workbooks.Open(report_bc)
    work_sheets = sheets.Worksheets[0]
    pdf=mmr_report.replace('19G','20')
    pdf=pdf.replace('Message MOU Revenue by BAN Report.pdf','Bill Count Trending.pdf')
    # Convert into PDF File
    work_sheets.ExportAsFixedFormat(0, pdf)
    sheets.Close()

def call(company,report_mps,FCCID=''):
    if company=='1':
        if(FCCID=='1'):
            Prepare_data(report_mps, 19,'25',client=2)
        if(FCCID=='2'):
            Prepare_data(report_mps, 5, '20',client=2)
        if(FCCID=='3'):
            Prepare_data(report_mps, 34, '25',client=2)   
    elif company=='2':
        Prepare_data(report_mps, 2,'22','9627_OR_','BendTel_Trending_','Sheet1',client=3)
    elif company=='3':
        Prepare_data(report_mps, 18,'21','GCI_SW_','_GCI_Bill Count Trending.xlsx',client=1)
    elif company=='4':
        Prepare_data(report_mps, 18,'10','MTA_SW_','_MTA_Bill Count Trending.xlsx',client=1)
